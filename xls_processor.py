#!/usr/bin/env python3

import os
import sys
import re
import csv
import math

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

import openpyxl
from tokenizer import shunting_yard

import collections

import random
import string

#import h5py
#import numpy as np

#import apihelper

import stopwatch

import warnings
warnings.filterwarnings("ignore")


if sys.version_info < (3,):
  def is_string(s):
    return type(s) in [str, unicode]
else:
  def is_string(s):
    return type(s) == str


def randomword(length):
  return ''.join(random.choice(string.ascii_lowercase) for i in range(length))


def slugify(value):
  import unicodedata
  value = unicodedata.normalize('NFKD', value)
  value = re.sub('[^\w\s-]', '', value).strip().lower()
  return re.sub('[-\s]+', '-', value)


class ExcelLoader:
  def __init__(self, filename):
    self.workbook_normal = openpyxl.load_workbook(filename, data_only=False)
    #self.workbook_data = openpyxl.load_workbook(filename, data_only=True)

  def iter_rows(self, sheet_name, data_only=True):
    #wb = data_only and self.workbook_data or self.workbook_normal
    wb = self.workbook_normal
    sheet = wb[sheet_name]
    return sheet.iter_rows()

  def iter_icells(self, sheet_name, data_only=True):
    """Return generator for all cells with offsets prepended"""
    #wb = data_only and self.workbook_data or self.workbook_normal
    wb = self.workbook_normal
    sheet = wb[sheet_name]

    for i, row in enumerate(sheet.iter_rows()):
      for j, cell in enumerate(row):
        if not cell:
          continue
        yield (i, j, cell)

  def sheet_names(self):
    return self.workbook_normal.get_sheet_names()


class ExcelFormula:
  @staticmethod
  def all_cell_coordinates(formula):
    rpn = shunting_yard(formula)
    for node in rpn:
      if node.token.ttype == 'operand' and node.token.tsubtype == 'range':
        yield node.token.tvalue

  @staticmethod
  def expand_cell_range(cell_range):
    """Return generator for offsets (tuples) of expanded cell range"""
    m = re.match(r'([A-Z]+[0-9]+):([A-Z]+[0-9]+)', cell_range)
    a = ExcelFormula.offsets_from_coordinates(m.group(1))
    b = ExcelFormula.offsets_from_coordinates(m.group(2))

    for i in range(a[0], b[0]+1):
      for j in range(a[1], b[1]+1):
        yield (i,j)

  @staticmethod
  def offsets_from_coordinates(coordinates):
    xy = openpyxl.utils.coordinate_from_string(coordinates) 
    col = openpyxl.utils.column_index_from_string(xy[0])
    row = xy[1] 
    return (row-1,col-1)


class CellData:
  Data = Qt.UserRole+1
  Expression = Qt.UserRole+2
  Coordinate = Qt.UserRole+3
  Category = Qt.UserRole+4
  SetCategory = Qt.UserRole+5
  Border = Qt.UserRole+6

class CellCategory:
  Empty = 0
  Label = 1
  Input = 2
  Output = 3
  Intermediate = 4
  Ignored = 5

  Any = 6

  @classmethod
  def desc(self, category):
    return {
      self.Empty: 'Empty',
      self.Label: 'Label',
      self.Input: 'Input',
      self.Output: 'Output',
      self.Intermediate: 'Intermediate',
      self.Ignored: 'Ignored'
    }[category]

  @classmethod
  def color(self, category):
    return {
      self.Empty: QColor(255,255,255),
      self.Label: QColor(200,200,200),
      self.Input: QColor(255,255,200),
      self.Output: QColor(200,255,200),
      self.Intermediate: QColor(255,200,200),
      self.Ignored: QColor(220,220,250)
    }[category]

  @classmethod
  def categories(self):
    return [
      self.Empty,
      self.Label,
      self.Input,
      self.Output,
      self.Intermediate,
      self.Ignored
    ]


class CellBorder:
  NoBorder = 0
  Top = 1
  Left = 2
  Right = 4
  Bottom = 8
  All = Top | Left | Right | Bottom

class Direction:
  Left = 0
  Right = 1
  Top = 2
  Bottom = 3


class Block:
  def __init__(self, model, top_left=(0,0), bottom_right=(0,0), type_=CellCategory.Label):
    self.top_left = top_left
    self.bottom_right = bottom_right
    self.type_ = type_
    self.model = model

    #print('{0}: {1}'.format(self.top_left, self.get_data(*self.top_left)))

  def get_data(self, row, column):
    return self.model.index(row + self.top_left[0], column + self.top_left[1]).data(CellData.Data)

  def dimensions(self):
    (ax, ay), (bx, by) = self.top_left, self.bottom_right
    return (bx-ax+1, by-ay+1)

  def row(self):
    return self.top_left[0]

  def column(self):
    return self.top_left[1]

  def row_count(self):
    return self.dimensions()[0]

  def column_count(self):
    return self.dimensions()[1]


class Blocks:
  def __init__(self):
    self.blocks = []

  def clear(self):
    del self.blocks[:]

  def add_block(self, block):
    self.blocks.append(block)

  def indices(self):
    return self.blocks

  def data_blocks(self):
    for  block in self.blocks:
      if block.type_ in [CellCategory.Input, CellCategory.Output, CellCategory.Intermediate]:
        yield block

  def next_block(self, block, direction, position=(-1,-1), dimensions=(0,0), type_=CellCategory.Any):
    matching_blocks = []

    for bl in self.blocks:
      if type_ != CellCategory.Any and bl.type_ != type_:
        continue

      if (dimensions[0] != 0 and dimensions[0] != bl.row_count()) or \
        (dimensions[1] != 0 and dimensions[1] != bl.column_count()):
        continue

      if (position[0] != -1 and position[0] != bl.row()) or \
        (position[1] != -1 and position[1] != bl.column()):
        continue

      if (direction == Direction.Top and block.row() <= bl.row()) or \
        (direction == Direction.Bottom and block.row() >= bl.row()) or \
        (direction == Direction.Left and block.column() <= bl.column()) or \
        (direction == Direction.Right and block.column() >= bl.column()):
        continue

      row_d, col_d = abs(block.row()-bl.row()), abs(block.column()-bl.column())
      distance = math.sqrt(pow(row_d, 2) + pow(col_d, 2))
      matching_blocks.append((distance, bl))

    sorted_blocks = sorted(matching_blocks, key=lambda x: x[0])

    if sorted_blocks:
      return sorted_blocks[0][1]
    else:
      return None


  def export(self, filename):
    for block in self.data_blocks():
      label_block = self.next_block(
        block, Direction.Left, (-1, -1), (block.row_count(), 1), CellCategory.Label
      )
      index_block = self.next_block(
        block, Direction.Top, (-1, -1), (0, block.column_count()), CellCategory.Label
      )
  
      if not (label_block and index_block):
        continue

      title_block = self.next_block(
        block, Direction.Top, (-1,-1), (0,0),  CellCategory.Any
      )
      if title_block.dimensions() == (1,1) and title_block.type_ == CellCategory.Label:
        title = title_block.get_data(0,0)
      else:
        title_block = self.next_block(
          block, Direction.Left, (-1,-1), (0,0), CellCategory.Any
        )
        if title_block.dimensions() == (1,1) and title_block.type_ == CellCategory.Label:
          title = title_block.get_data(0,0)
        else:
          title = randomword(10)
           
      filename = slugify(title) + '.csv'

      os.chdir('tmp')
      with open(filename, 'w', newline='') as f:
        a = csv.writer(f)

        index_data = []
        rows, columns = index_block.dimensions()
        for i in range(columns):
          index_data.append(index_block.get_data(rows-1,i))

        row_data = ['Index']
        w,h = label_block.dimensions()
        for i in range(w):
          for j in range(h):
            row_data.append(label_block.get_data(i,j))
        a.writerow(row_data)

        rows, columns = block.dimensions()
        for i in range(columns):
          row_data = [index_data[i]]
          for j in range(rows):
            row_data.append(block.get_data(j,i))
          a.writerow(row_data)

      os.chdir('..')


class SheetModel(QStandardItemModel):
  def __init__(self, sheet_name, excel_loader, parent=None):
    super(SheetModel, self).__init__(parent)

    self.excel_loader = excel_loader
    self.sheet_name = sheet_name

    self.blocks = Blocks()

    for row in self.excel_loader.iter_rows(sheet_name, data_only=False):
      row_items = []
      for cell in row:

        if cell.value != None:
          if is_string(cell.value):
            text = cell.value.strip()  
          else:
            text = str(cell.value)
        else:
          text = ''

        item = QStandardItem(text)

        item.setData(cell.value, CellData.Data)
        item.setData(CellBorder.NoBorder, CellData.Border)
        item.setData(-1, CellData.SetCategory)
        item.setData(
          text and CellCategory.Label or CellCategory.Empty,
          CellData.Category
        )

        if cell.formula:
          item.setData(cell.formula, CellData.Expression)

        row_items.append(item)

      self.appendRow(row_items)


  def calculate_references(self):
    """Return dict (sheet_name -> cell references to that sheet)"""
    refs = {}
    for i in range(self.rowCount()):
      for j in range(self.columnCount()):
        text = self.data(self.index(i,j), CellData.Expression)
        if not text:
          continue

        if is_string(text) and text.startswith('='):
          self.setData(self.index(i,j), CellCategory.Output, CellData.Category)

          cell_coordinates = ExcelFormula.all_cell_coordinates(text)

          for cc in cell_coordinates:
              #print('{2} ({0},{1}): {3}'.format(i,j,self.sheet_name, cc))
              m = re.match(r'(([^!]+)\!)?([A-Z]+[0-9]+)(:[A-Z]+[0-9]+)?', cc)

              if not m:
                # cell is error!
                self.setData(self.index(i,j), '\''+text, CellData.Expression)
                self.setData(self.index(i,j), '\''+text, CellData.Data)
                self.item(i,j).setText(text)
                break

              sheet_name = m.group(2) or self.sheet_name
              ref = m.group(3) + (m.group(4) or '')

              if not sheet_name in refs: refs[sheet_name] = []
              refs[sheet_name].append(ref)

    return refs

  def apply_references(self, refs):
    for ref in refs:
      if ':' in ref:
        indices = [self.index(*ret) for ret in ExcelFormula.expand_cell_range(ref)]
      else:
        indices = [self.index(*ExcelFormula.offsets_from_coordinates(ref))]

      for index in indices:
        if not index:
          continue
        current_category = self.data(index, CellData.Category)
        if current_category == CellCategory.Output:
          self.setData(index, CellCategory.Intermediate, CellData.Category)
        elif self.data(index, CellData.Category) == CellCategory.Label:
          self.setData(index, CellCategory.Input, CellData.Category)

  def update(self):
    #self.blocks = list(self.scan_blocks())
    self.blocks.clear()
    [self.blocks.add_block(b) for b in self.scan_blocks()]

  def reset_categories(self):
    for col in range(self.columnCount()):
      for row in range(self.rowCount()):
        self.setData(self.index(row, col), -1, CellData.SetCategory)
    self.update()

  def scan_blocks(self):
    def row_sections(row):
      start = 0
      for x in range(self.columnCount()):
        #cat = self.data(self.index(row, x), CellData.Category)
        cat = self.get_category(self.index(row, x))

        x_next = x+1

        if cat == CellCategory.Empty:
          start = x_next
          continue

        if (x_next>self.columnCount()-1 or
             #self.data(self.index(row, x_next), CellData.Category) != cat):
             self.get_category(self.index(row, x_next)) != cat):
          yield (start, x)
          start = x_next

    all_rows = []
    for i in range(self.rowCount()):
      all_rows.append(list(row_sections(i)))
      
    blacklist = {}

    for i, row in enumerate(all_rows):
      for sec in row:
        blacklist = {b: blacklist[b] for b in blacklist if i<=blacklist[b]}

        if sec in blacklist:
          continue

        k = i
        while True:
          if k+1 >= self.rowCount(): 
            break

          flag = False
          for m in all_rows[k+1]:
            if m==sec:
              flag = self.get_category(self.index(k+1, m[0])) == \
                     self.get_category(self.index(i, sec[0]))
#              flag = self.index(k+1,m[0]).data(CellData.Category) == \
#                    self.index(i,sec[0]).data(CellData.Category)
              break
          if not flag:
            break

          k += 1
          blacklist[sec] = k

        #cat = self.data(self.index(i, sec[0]), CellData.Category)
        cat = self.get_category(self.index(i, sec[0]))
        yield Block(self, (i, sec[0]), (k, sec[1]), cat)

  def get_category(self, index):
    ret = self.data(index, CellData.SetCategory)
    if ret == -1:
      ret = self.data(index, CellData.Category)
    return ret

  def data(self, index, role):
    if role == Qt.BackgroundRole:
      category = self.get_category(index)

      if category == CellCategory.Input:
        return QColor(255,255,200)
      elif category == CellCategory.Intermediate:
        return QColor(255,200,200)
      elif category == CellCategory.Output:
        return QColor(200,255,200)
      elif category == CellCategory.Label:
        return QColor(200,200,200)
      elif category == CellCategory.Ignored:
        return QColor(220,220,250)
      return QColor(255,255,255)

    if role == CellData.Border:
      border = CellBorder.NoBorder
      x,y = index.row(), index.column()

      for block in self.blocks.indices():
        (ax, ay), (bx, by) = block.top_left, block.bottom_right
        if x>=ax and x<=bx:
          if y==ay:
            border |= CellBorder.Left   
          if y==by:
            border |= CellBorder.Right
        if y>=ay and y<=by:
          if x==ax:
            border |= CellBorder.Top
          if x==bx:
            border |= CellBorder.Bottom
      return border
   
    return super(SheetModel, self).data(index, role)


class WorkbookModel:
  def load_file(self, filename):
    self.excel_loader = ExcelLoader(filename)

    sheet_names = self.excel_loader.sheet_names()
    self.sheet_models = {}

    if not sheet_names:
      return

    for sheet_name in sheet_names:
      self.sheet_models[sheet_name] = SheetModel(sheet_name, self.excel_loader)

    for sheet in self.sheet_models:
      refs = self.sheet_models[sheet].calculate_references()
      for ref in refs:
        self.sheet_models[ref].apply_references(refs[ref])

    for sheet in self.sheet_models:
      self.sheet_models[sheet].update()

    self.current_sheet_name = sheet_names[0]

  def set_sheet_by_index(self, index):
    sheet_names = self.excel_loader.sheet_names()
    self.current_sheet_name = sheet_names[index]

  def sheet_names(self):
    return self.excel_loader.sheet_names()

  def current_sheet_model(self):
    return self.sheet_models[self.current_sheet_name]

def intensify(qcolor):
  hsv = qcolor.getHsv()
  brightness = hsv[0] >= 0 and 255 or hsv[2]-100
  return QColor.fromHsv(hsv[0], 255, brightness)

class ItemDelegate(QStyledItemDelegate):
  def __init__(self, parent=None):
    super(ItemDelegate, self).__init__(parent)

  def paint(self, painter, option, index):
    super(ItemDelegate, self).paint(painter, option, index)

    thickness = 3

#    color = index.data(Qt.BackgroundRole).getHsv()
#    brightness = color[0] >= 0 and 255 or 100
#    color = QColor.fromHsv(color[0], 255, brightness)
    color = intensify(index.data(Qt.BackgroundRole))

    pen = QPen()
    pen.setColor(color)
    pen.setWidth(thickness)
    painter.setPen(pen)

    border = index.data(CellData.Border)

    thickness /= 2.0
    rect = option.rect.adjusted(thickness, thickness, -thickness, -thickness)
  
    if border & CellBorder.Top:
      painter.drawLine(rect.topLeft(), rect.topRight())
    if border & CellBorder.Left:
      painter.drawLine(rect.topLeft(), rect.bottomLeft())
    if border & CellBorder.Right:
      painter.drawLine(rect.topRight(), rect.bottomRight())
    if border & CellBorder.Bottom:
      painter.drawLine(rect.bottomLeft(), rect.bottomRight())

  def editorEvent(self, event, model, option, index):
    return False

class TableView(QTableView):
  hovered_cell_changed = pyqtSignal(int,int)
  pointer_left = pyqtSignal()

  def __init__(self, parent=None):
    super(TableView, self).__init__(parent)
    self.setItemDelegate(ItemDelegate(self))
    self.setMouseTracking(True)
    self.viewport().setAttribute(Qt.WA_Hover, True)

    self.hovered_cell = None

    self.installEventFilter(self)

    self.build_menu()

  def build_menu(self):
    self.menu = QMenu('', self)
    ag = QActionGroup(self.menu)

    ag.triggered[QAction].connect(self.context_action_triggered)

    actions = {
      CellCategory.Label: 'Label',
      CellCategory.Input: 'Input',
      CellCategory.Intermediate: 'Intermediate',
      CellCategory.Output: 'Output',
      CellCategory.Ignored: 'Ignored'
    }

    for a in actions:
      qa = QAction(actions[a], ag)
      qa.setData(a)
      qa.setCheckable(True)
      self.menu.addAction(qa)
      
    self.setContextMenuPolicy(Qt.CustomContextMenu)
    self.customContextMenuRequested[QPoint].connect(self.show_context_menu)

  def show_context_menu(self, qpoint):
    selmodel = self.selectionModel()
    if not selmodel:
      return
    indexes = selmodel.selection().indexes()
    if not indexes:
      return
    #cats = [ix.data(CellData.Category) for ix in indexes]
    cats = [self.model().get_category(ix) for ix in indexes]
    for ac in self.menu.actions():
      if cats.count(cats[0]) == len(cats) and ac.data() == cats[0]:
        ac.setChecked(True)
      else:
        ac.setChecked(False)

    self.menu.popup(self.viewport().mapToGlobal(qpoint))

  def context_action_triggered(self, qaction):
    selmodel = self.selectionModel()
    if not selmodel:
      return
    indexes = selmodel.selection().indexes()
    if not indexes:
      return
    for ix in indexes:
      #if ix.data(CellData.Category) != CellCategory.Empty:
      self.model().setData(ix, qaction.data(), CellData.SetCategory)
    self.model().update()
  

  def update_hovered(self, pos):
    index = self.indexAt(pos)
    if index != self.hovered_cell:
      self.hovered_cell = index
      self.hovered_cell_changed.emit(index.row(),index.column())

  def mouseMoveEvent(self, event):
    self.update_hovered(event.pos())
    super(TableView, self).mouseMoveEvent(event)

  def wheelEvent(self, event):
    self.update_hovered(event.pos())
    super(TableView, self).wheelEvent(event)

  def eventFilter(self, obj, event):
    if obj==self and event.type()==QEvent.Leave:
      self.pointer_left.emit()
      self.hovered_cell = None
    return super(TableView, self).eventFilter(obj, event)

  def do(self, what):
    model = self.model()

    selmodel = self.selectionModel()
    if not selmodel:
      return
    indexes = selmodel.selection().indexes()
    if not indexes:
      return

    x,y = indexes[0].column(), indexes[0].row()
    xmin,xmax,ymin,ymax = x,x,y,y
    for ix in indexes:
      xmin = min(xmin, ix.column())
      xmax = max(xmax, ix.column())
      ymin = min(ymin, ix.row())
      ymax = max(ymax, ix.row())
    rows = ymax - ymin + 1
    columns = xmax - xmin + 1

    #print('{0} {1} {2} {3}'.format(xmin, xmax, ymin, ymax))
#      #if ix.data(CellData.Category) != CellCategory.Empty:
#      self.table.model().setData(ix, qaction.data(), CellData.SetCategory)
#    self.table.model().update()

    if what == 'merge row-wise':
      pass
    elif what == 'merge column-wise':
      pass
    elif what == 'add row':
      if rows == 1:
        self.model().insertRow(y+1)
    elif what == 'add column':
      pass
    elif what == 'delete row':
      for i in range(ymin, ymax+1):
        self.model().takeRow(i)
    elif what == 'delete column':
      pass

class MyToolBar(QToolBar):
  def __init__(self, parent=None):
    super(MyToolBar, self).__init__(parent)
    self.setMovable(False)

  def add_action(self, data, pic, label, parent, func=None, shortcut=None, checkable=False, checked=False):
    a = QAction(QIcon('pics/'+pic), label, parent)

    a.setData(data)

    if shortcut:
      a.setShortcut(shortcut)

    if checkable:
      a.setCheckable(True)
      a.setChecked(checked)

    if func:
      a.triggered.connect(func)

    self.addAction(a)

  def add_group(self, parent, func=None):
    ag = QActionGroup(parent)

    if func:
      ag.triggered[QAction].connect(func)

    return ag

  def add_stretch(self):
    w = QWidget()
    w.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
    self.addWidget(w)

class MainToolBar(MyToolBar):
  def __init__(self, parent=None):
    super(MainToolBar, self).__init__(parent)

    self.add_action('open', 'open.png', '&Open File', parent, parent.open_file, 'Ctrl-O')

    self.addSeparator()

    ag = self.add_group(self, parent.mode_changed)
    self.add_action('categories', 'categories.png', 'Category Manipulator', ag, checkable=True, checked=True)
    self.add_action('table', 'table.png', 'Table Manipulators', ag, checkable=True)
    self.add_action('hdf5', 'hdf5.png', 'HDF5 Preview', ag, checkable=True)

    self.add_stretch()

    self.add_action('exit', 'exit.png', 'E&xit', parent, qApp.quit, 'Ctrl-Q')

class CategoryToolBar:
  def __init__(self, toolbar):
    self.toolbar = toolbar

    self.t2 = MyToolBar()
    self.t2.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

    self.toolbar.add_action('reset', 'reset.png', '&Reset categories', toolbar.parent,
        toolbar.parent.reset_categories)

    toolbar.parent.addToolBar(self.t2)
    ag = self.t2.add_group(toolbar.parent, toolbar.parent.set_category)

    self.t2.add_action(CellCategory.Empty, 'category_white.png', 'Empty', ag)
    self.t2.add_action(CellCategory.Label, 'category_grey.png', 'Label', ag)
    self.t2.add_action(CellCategory.Input, 'category_yellow.png', 'Input', ag)
    self.t2.add_action(CellCategory.Output, 'category_green.png', 'Output', ag)
    self.t2.add_action(CellCategory.Intermediate, 'category_red.png', 'Intermediate', ag)
    self.t2.add_action(CellCategory.Ignored, 'category_blue.png', 'Ignored', ag)

  def __del__(self):
    self.toolbar.parent.removeToolBar(self.t2)
    self.toolbar.clear()

class TableToolBar:
  def __init__(self, toolbar):
    self.toolbar = toolbar
    
    ag = toolbar.add_group(toolbar.parent, toolbar.parent.table_action)
    toolbar.add_action('merge row-wise', 'merge1.png', 'Merge selected contents row-wise', ag)
    toolbar.add_action('merge column-wise', 'merge2.png', 'Merge selected contents column-wise', ag)
    toolbar.add_action('add row', 'add_row.png', 'Add row', ag)
    toolbar.add_action('add column', 'add_column.png', 'Add column', ag)
    toolbar.add_action('delete row', 'delete_row.png', 'Delete row', ag)
    toolbar.add_action('delete column', 'delete_column.png', 'Delete column', ag)

  def __del__(self):
    self.toolbar.clear()

class HDF5ToolBar:
  def __init__(self, toolbar):
    self.toolbar = toolbar
    toolbar.add_action('export', 'export.png', '&Export', toolbar.parent, toolbar.parent.export)

  def __del__(self):
    self.toolbar.clear()

class SubToolBar(MyToolBar):
  def __init__(self, parent=None):
    super(SubToolBar, self).__init__(parent)
    self.parent = parent

  def load(self, what):
    try:
      del self.toolbar
    except AttributeError:
      pass

    if what == 'categories':
      self.toolbar = CategoryToolBar(self)
    elif what == 'table':
      self.toolbar = TableToolBar(self)
    elif what == 'hdf5':
     self.toolbar = HDF5ToolBar(self)

  def update(self):
    pass

class MainWindow(QMainWindow):
  def __init__(self, parent=None):
    super(MainWindow, self).__init__(parent)
    self.setWindowTitle('XLS Processor')
 
    self.workbook_model = WorkbookModel()
    self.setupUI()

  def setupUI(self):
    vbox = QVBoxLayout()

    self.addToolBar(MainToolBar(self))
    self.addToolBarBreak()
    self.subtoolbar = SubToolBar(self)
    self.addToolBar(self.subtoolbar)
    self.subtoolbar.load('categories')

    #toolbar2 = self.addToolBar('toolbar2')
    #toolbar2.setMovable(False)
    #resetAction = QAction(QIcon('pics/reset.png'), '&Reset categories', self)
    #resetAction.triggered.connect(self.reset_categories)
    #toolbar2.addAction(resetAction)

    #menubar = self.menuBar()
    #fileMenu = menubar.addMenu('&File')
    #fileMenu.addAction(openFileAction)
    #fileMenu.addAction(exportAction)
    #fileMenu.addSeparator()
    #fileMenu.addAction(exitAction)

    self.statusLabel, self.cellinfoLabel = QLabel(), QLabel()
    self.statusLabel.setFrameStyle(QFrame.Panel | QFrame.Sunken)
    self.cellinfoLabel.setFrameStyle(QFrame.Panel | QFrame.Sunken)

    self.statusBar().addPermanentWidget(self.statusLabel, 2)
    self.statusBar().addPermanentWidget(self.cellinfoLabel, 1)

    colordict = collections.OrderedDict()
    for cat in CellCategory.categories():
      colordict[CellCategory.desc(cat)] = CellCategory.color(cat)

    self.table = TableView()
    self.table.hovered_cell_changed[int,int].connect(self.update_cellinfo)
    self.table.pointer_left.connect(lambda: self.cellinfoLabel.setText(''))
    vbox.addWidget(self.table)

    self.tabs = QTabBar()
    self.tabs.setShape(QTabBar.TriangularSouth)
    self.tabs.setExpanding(False)
    self.tabs.currentChanged[int].connect(self.tab_changed)
    vbox.addWidget(self.tabs)

    vbox.setSpacing(0)

    widget = QWidget()
    widget.setLayout(vbox)
    self.setCentralWidget(widget)

  def mode_changed(self, qaction):
    self.subtoolbar.load(qaction.data())
 
  def reset_categories(self):
    self.workbook_model.current_sheet_model().reset_categories()

  def set_category(self, qaction):
    selmodel = self.table.selectionModel()
    if not selmodel:
      return
    indexes = selmodel.selection().indexes()
    if not indexes:
      return
    for ix in indexes:
      #if ix.data(CellData.Category) != CellCategory.Empty:
      self.table.model().setData(ix, qaction.data(), CellData.SetCategory)
    self.table.model().update()

  def table_action(self, qaction):
    self.table.do(qaction.data())

  def export(self):
    self.table.model().blocks.export('out.hd5')

  def status_message(self, msg, timeout=4000):
    self.statusLabel.setText(msg)
    if timeout:
      QTimer.singleShot(timeout, lambda: self.statusLabel.setText(''))

  def update_cellinfo(self, row, column):
    if -1 in [row, column]:
      return

    #category = self.table.model().index(row,column).data(CellData.Category)
    model = self.table.model()
    category = model.get_category(model.index(row,column))
    text = '({0},{1})'.format(row+1, column+1)
    if category and category != CellCategory.Empty:
      text += ' {0}'.format(CellCategory.desc(category))
    self.cellinfoLabel.setText(text)

  def tab_changed(self, index):
    self.workbook_model.set_sheet_by_index(index)
    self.update_table()

  def update_tabbar(self):
    while self.tabs.count()>0:
      self.tabs.removeTab(0)

    for name in self.workbook_model.sheet_names():
      self.tabs.addTab(name.replace('&','&&'))

  def update_table(self):
    self.table.setModel(self.workbook_model.current_sheet_model())

  def load_workbook(self, filename):
    self.status_message('Loading file \'{0}\'...'.format(filename), timeout=0)
    QApplication.processEvents()
    self.workbook_model.load_file(filename)
    self.update_tabbar()
    self.status_message('\'{0}\' loaded.'.format(filename))


  def open_file(self):
    settings = QSettings()
    dialog = QFileDialog(self, 'Select Excel-File', str(settings.value('default_dir')), 'Excel-Files (*.xls *.xlsx)')
    dialog.setFileMode(QFileDialog.ExistingFile)
    dialog.setOption(QFileDialog.DontUseNativeDialog)
    filename = ''
    if dialog.exec_():
      filename = dialog.selectedFiles()[0]
    if not filename:
      return
    directory = QDir().absoluteFilePath(filename)
    settings.setValue('default_dir', directory)
    self.load_workbook(filename)

    

if __name__ == '__main__':
  if len(sys.argv) > 2:
    sys.exit('Usage: {0} [filename]'.format(sys.argv[0]))

  app = QApplication(sys.argv)
  app.aboutToQuit.connect(app.deleteLater)
  app.setApplicationName('XLS Processor')
  app.setOrganizationName('Excel Master')
  app.setOrganizationDomain('excel-master.com')

  w = MainWindow()
  w.resize(1200,800)
  w.setWindowIcon(QIcon('pics/icon.png'))

  t = stopwatch.Timer()

  if len(sys.argv) == 2:
    w.load_workbook(sys.argv[1])

  t.stop()
  print(t.elapsed)

  w.show()
  app.exec_()
  sys.exit()
